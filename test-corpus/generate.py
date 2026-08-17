#!/usr/bin/env python3
"""Generate a synthetic .xlsx test corpus for the Marksheet XLSX importer/exporter.

Why generated rather than downloaded: the project's own conversion fixtures
(tests/conversion/README.md) deliberately avoid checking in copied Office
files, generating deterministic OOXML instead. This script follows the same
spirit -- it is the checked-in, reproducible artifact; the binary .xlsx files
it writes to xlsx/ are gitignored build output. Regenerate with:

    python3 -m venv .venv && . .venv/bin/activate
    pip install -r requirements.txt
    python3 generate.py

Every file is annotated in manifest.json (tier, description, notable
features) which README.md is generated from.
"""
from __future__ import annotations

import json
import shutil
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo

from functions_catalog import CATALOG

HERE = Path(__file__).resolve().parent
OUT_DIR = HERE / "xlsx"

MANIFEST: list[dict] = []

PORTABLE_A1_FUNCTIONS = {
    "SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA",
    "IF", "AND", "OR", "NOT", "IFERROR",
    "ABS", "ROUND", "ROUNDUP", "ROUNDDOWN", "INT", "MOD",
    "CONCAT", "LEFT", "RIGHT", "MID", "LEN", "LOWER", "UPPER", "TRIM",
    "INDEX", "MATCH",
    "DATE", "YEAR", "MONTH", "DAY",
    "ISBLANK", "ISNUMBER", "ISTEXT", "ISERROR",
}

BOLD_HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)
MONEY = '#,##0.00_);[Red](#,##0.00)'
CURRENCY_USD = '"$"#,##0.00'
PERCENT = "0.0%"


def out(name: str) -> Path:
    return OUT_DIR / name


def style_header_row(ws, row=1, last_col=1):
    for col in range(1, last_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = BOLD_HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def record(filename: str, tier: str, title: str, description: str, features: list[str]):
    MANIFEST.append(
        {
            "file": filename,
            "tier": tier,
            "title": title,
            "description": description,
            "features": features,
        }
    )


# --------------------------------------------------------------------------
# Raw-zip helpers: inject OOXML parts openpyxl cannot author itself, so we
# can exercise Marksheet's Macro / PivotTable / ExternalLink detection paths
# (crates/marksheet-convert/src/xlsx/import.rs matches on part-name prefixes
# such as "xl/pivot", "xl/externalLinks/", and the workbook's declared
# content type for macro detection -- see record_unconsumed_package_content).
# --------------------------------------------------------------------------

def add_raw_zip_entries(xlsx_path: Path, entries: dict[str, bytes]) -> None:
    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as src, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            dst.writestr(item, src.read(item.filename))
        for name, content in entries.items():
            dst.writestr(name, content)
    tmp.replace(xlsx_path)


def rewrite_zip_entry(xlsx_path: Path, name: str, transform) -> None:
    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as src, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            content = src.read(item.filename)
            if item.filename == name:
                content = transform(content)
            dst.writestr(item, content)
    tmp.replace(xlsx_path)


def make_macro_enabled(xlsx_path: Path, xlsm_path: Path) -> None:
    """Flip the workbook content type to macroEnabled.main and add a VBA stub.

    Mirrors crates/marksheet-convert/src/xlsx/import.rs's own test fixture
    (unconsumed_parts_relationships_and_macro_mime_are_lossy), which proves a
    placeholder byte string for xl/vbaProject.bin is sufficient to exercise
    the Macro-feature path -- real VBA bytecode is not required.
    """
    def flip(content: bytes) -> bytes:
        return content.replace(
            b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
        )

    rewrite_zip_entry(xlsx_path, "[Content_Types].xml", flip)
    add_raw_zip_entries(
        xlsx_path,
        {"xl/vbaProject.bin": b"placeholder VBA project -- not real bytecode, see README.md"},
    )
    shutil.move(str(xlsx_path), str(xlsm_path))


PIVOT_CACHE_DEFINITION = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  r:id="rId1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
  refreshOnLoad="1" recordCount="10">
  <cacheSource type="worksheet"><worksheetSource ref="A1:D11" sheet="Transactions"/></cacheSource>
  <cacheFields count="2">
    <cacheField name="Region" numFmtId="0"><sharedItems/></cacheField>
    <cacheField name="Amount" numFmtId="0"><sharedItems containsSemiMixedTypes="0" containsString="0" containsNumber="1"/></cacheField>
  </cacheFields>
</pivotCacheDefinition>
"""

PIVOT_TABLE_XML = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  name="RegionPivot" cacheId="1" applyNumberFormats="0" applyBorderFormats="0"
  applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1"
  dataCaption="Values" updatedVersion="6" minRefreshableVersion="3" useAutoFormatting="1"
  itemPrintTitles="1" createdVersion="6" indent="0" outline="1" outlineData="1" multipleFieldFilters="0">
  <location ref="A3:B6" firstHeaderRow="1" firstDataRow="2" firstDataCol="1"/>
  <pivotFields count="2">
    <pivotField axis="axisRow" showAll="0"><items count="1"><item x="0"/></items></pivotField>
    <pivotField dataField="1" showAll="0"/>
  </pivotFields>
  <rowFields count="1"><field x="0"/></rowFields>
  <dataFields count="1"><dataField name="Sum of Amount" fld="1" baseField="0" baseItem="0"/></dataFields>
</pivotTableDefinition>
"""

EXTERNAL_LINK_XML = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <externalBook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1">
    <sheetNames><sheetName val="Sheet1"/></sheetNames>
    <sheetDataSet>
      <sheetData sheetId="0">
        <row r="1"><cell r="A1"><v>123</v></cell></row>
      </sheetData>
    </sheetDataSet>
  </externalBook>
</externalLink>
"""


def add_pivot_table_stub(xlsx_path: Path) -> None:
    """Minimal, deliberately-not-fully-wired pivot table parts.

    marksheet-convert only needs to *see* a part under an "xl/pivot" prefix
    to record a PivotTable/Unsupported outcome (see import.rs), so this stub
    does not need a real relationship graph to exercise that path. It is not
    a substitute for a real Excel-authored pivot table -- see README.md.
    """
    add_raw_zip_entries(
        xlsx_path,
        {
            "xl/pivotCache/pivotCacheDefinition1.xml": PIVOT_CACHE_DEFINITION,
            "xl/pivotTables/pivotTable1.xml": PIVOT_TABLE_XML,
        },
    )


def add_external_link_stub(xlsx_path: Path) -> None:
    add_raw_zip_entries(xlsx_path, {"xl/externalLinks/externalLink1.xml": EXTERNAL_LINK_XML})


# --------------------------------------------------------------------------
# openpyxl consistently writes package-*absolute* relationship targets
# (Target="/xl/worksheets/sheet1.xml") for worksheets, tables, charts, and
# comments -- valid under the OPC spec, but crates/marksheet-convert/src/
# xlsx/package.rs's resolve_target() deliberately rejects any target
# starting with "/" as "absolute, external, or malformed" (see MS4105). That
# makes every openpyxl-authored file unimportable as-is. We normalize targets
# to the relative form real Excel emits so the rest of this corpus can
# actually exercise formula/feature handling -- see
# 31_openpyxl_absolute_relationship_targets.xlsx for a deliberately
# unnormalized fixture that documents the gap instead of hiding it.
# --------------------------------------------------------------------------

import posixpath
import re


def _source_part_for_rels(rels_part: str) -> str:
    directory, filename = posixpath.split(rels_part)
    base = filename[: -len(".rels")]
    parent = posixpath.dirname(directory)
    return posixpath.join(parent, base) if parent else base


def normalize_absolute_rels(xlsx_path: Path) -> None:
    def transform(name: str, content: bytes) -> bytes:
        if not name.endswith(".rels"):
            return content
        source_dir = posixpath.dirname(_source_part_for_rels(name))
        text = content.decode("utf-8")

        def repl(match: re.Match) -> str:
            target = match.group(1)
            rel = posixpath.relpath(target[1:], source_dir or ".")
            return f'Target="{rel}"'

        return re.sub(r'Target="(/[^"]*)"', repl, text).encode("utf-8")

    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as src, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            dst.writestr(item, transform(item.filename, src.read(item.filename)))
    tmp.replace(xlsx_path)


# --------------------------------------------------------------------------
# Showcase-sheet seed data (columns H:N), reused by every category sheet in
# 30_formula_function_showcase.xlsx.
# --------------------------------------------------------------------------

SEED_NUMBERS = [5, -3, 12, 7.5, 0, 100, -22.25, 3, 9, 45]
SEED_NUMBERS2 = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20]
SEED_TEXT = ["Alpha", "beta", "Gamma", "delta", "EPSILON", "zeta", "Eta", "theta", "iota", "kappa"]


def write_seed_block(ws) -> None:
    headers = ["Num", "Num2", "Text", "Date", "Bool", "Key", "Value"]
    for offset, label in enumerate(headers):
        ws.cell(row=1, column=8 + offset, value=label).font = Font(bold=True)
    for i in range(10):
        r = i + 2
        ws.cell(row=r, column=8, value=SEED_NUMBERS[i])
        ws.cell(row=r, column=9, value=SEED_NUMBERS2[i])
        ws.cell(row=r, column=10, value=SEED_TEXT[i])
        dcell = ws.cell(row=r, column=11, value=date(2024, 1, 1) + timedelta(days=i))
        dcell.number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=12, value=(i % 2 == 0))
        ws.cell(row=r, column=13, value=chr(65 + i))
        ws.cell(row=r, column=14, value=(i + 1) * 10)
    for col in "HIJKLMN":
        ws.column_dimensions[col].width = 11


# ==========================================================================
# TIER 1 -- Minimal
# ==========================================================================

def build_01_empty_workbook():
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.save(out("01_empty_workbook.xlsx"))
    record(
        "01_empty_workbook.xlsx", "1-minimal", "Empty workbook",
        "One sheet, zero cells. Lower bound for a workbook the importer must still accept cleanly.",
        ["empty sheet"],
    )


def build_02_single_cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 42
    wb.save(out("02_single_cell.xlsx"))
    record(
        "02_single_cell.xlsx", "1-minimal", "Single cell",
        "One sheet, one literal number in A1.",
        ["scalar cell"],
    )


def build_03_scalar_values_no_formulas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Values"
    ws.append(["Type", "Value"])
    rows = [
        ("Integer", 7),
        ("Negative", -13),
        ("Decimal", 3.14159),
        ("Large number", 123456789.5),
        ("Text", "hello world"),
        ("Empty string", ""),
        ("Boolean true", True),
        ("Boolean false", False),
        ("Date", date(2024, 6, 1)),
        ("Datetime", datetime(2024, 6, 1, 13, 45)),
        ("Currency-formatted", 1999.99),
        ("Percent-formatted", 0.42),
        ("Zero", 0),
    ]
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=value)
        if label == "Currency-formatted":
            c.number_format = CURRENCY_USD
        if label == "Percent-formatted":
            c.number_format = PERCENT
        if label == "Datetime":
            c.number_format = "yyyy-mm-dd hh:mm"
    style_header_row(ws, last_col=2)
    # Deliberate gap: leave row 20 entirely blank, then one more value past it.
    ws.cell(row=25, column=1, value="After a gap")
    ws.cell(row=25, column=2, value=1)
    wb.save(out("03_scalar_values_no_formulas.xlsx"))
    record(
        "03_scalar_values_no_formulas.xlsx", "1-minimal", "Scalar values, no formulas",
        "Every literal scalar type (int, negative, decimal, text, blank, bool, date, datetime, "
        "currency/percent number formats) with no formulas at all, plus a sparse gap before a "
        "trailing value.",
        ["scalar types", "number formats", "sparse rows"],
    )


# ==========================================================================
# TIER 2 -- Simple, everyday
# ==========================================================================

def build_04_personal_budget():
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws.append(["Category", "Budgeted", "Actual", "Difference"])
    items = [
        ("Rent", 1500, 1500), ("Utilities", 200, 224.50), ("Groceries", 400, 378.20),
        ("Transport", 150, 140), ("Entertainment", 100, 132.75), ("Savings", 500, 500),
        ("Insurance", 120, 120), ("Subscriptions", 45, 52.97), ("Misc", 80, 61.10),
    ]
    for i, (cat, budget, actual) in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=budget).number_format = CURRENCY_USD
        ws.cell(row=i, column=3, value=actual).number_format = CURRENCY_USD
        ws.cell(row=i, column=4, value=f"=B{i}-C{i}").number_format = CURRENCY_USD
    total_row = len(items) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    for col in "BCD":
        c = ws.cell(row=total_row, column=ord(col) - 64, value=f"=SUM({col}2:{col}{total_row - 1})")
        c.font = Font(bold=True)
        c.number_format = CURRENCY_USD
    style_header_row(ws, last_col=4)
    ws.column_dimensions["A"].width = 16
    wb.save(out("04_personal_budget.xlsx"))
    record(
        "04_personal_budget.xlsx", "2-simple", "Personal monthly budget",
        "Classic budget sheet: SUM totals, a per-row subtraction formula, currency formatting.",
        ["SUM", "subtraction formula", "currency format"],
    )


def build_05_grade_book():
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"
    ws.append(["Student", "Test 1", "Test 2", "Test 3", "Homework", "Average", "Letter Grade"])
    names = ["Amara", "Ben", "Carmen", "Deshawn", "Elif", "Farid", "Grace", "Hiro", "Ines", "Jamal", "Kira", "Liam"]
    import random
    random.seed(7)
    for i, name in enumerate(names, start=2):
        scores = [random.randint(55, 100) for _ in range(4)]
        ws.cell(row=i, column=1, value=name)
        for j, s in enumerate(scores, start=2):
            ws.cell(row=i, column=j, value=s)
        ws.cell(row=i, column=6, value=f"=AVERAGE(B{i}:E{i})").number_format = "0.0"
        ws.cell(
            row=i, column=7,
            value=(
                f'=IF(F{i}>=90,"A",IF(F{i}>=80,"B",IF(F{i}>=70,"C",'
                f'IF(F{i}>=60,"D","F"))))'
            ),
        )
    last = len(names) + 2
    ws.cell(row=last, column=1, value="Class stats").font = Font(bold=True)
    ws.cell(row=last, column=6, value=f"=AVERAGE(F2:F{last - 1})")
    ws.cell(row=last + 1, column=1, value="Highest average")
    ws.cell(row=last + 1, column=6, value=f"=MAX(F2:F{last - 1})")
    ws.cell(row=last + 2, column=1, value="Lowest average")
    ws.cell(row=last + 2, column=6, value=f"=MIN(F2:F{last - 1})")
    ws.cell(row=last + 3, column=1, value="Students counted")
    ws.cell(row=last + 3, column=6, value=f"=COUNT(F2:F{last - 1})")
    style_header_row(ws, last_col=7)
    wb.save(out("05_grade_book.xlsx"))
    record(
        "05_grade_book.xlsx", "2-simple", "Class grade book",
        "AVERAGE per student, nested IF letter grading, and MAX/MIN/COUNT class statistics -- "
        "every one of these functions is inside Marksheet's portable-a1@1 profile.",
        ["AVERAGE", "nested IF", "MAX", "MIN", "COUNT"],
    )


def build_06_todo_checklist():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(["Task", "Status", "Due Date", "Days Left", "Done"])
    tasks = [
        ("Draft proposal", "Done", date(2024, 1, 10)),
        ("Review budget", "In Progress", date(2024, 2, 1)),
        ("Send invoices", "Open", date(2024, 2, 15)),
        ("Team retro", "Done", date(2024, 1, 20)),
        ("Renew domain", "Open", date(2024, 3, 1)),
        ("Update handbook", "In Progress", date(2024, 2, 20)),
        ("Book travel", "Open", date(2024, 3, 10)),
        ("File taxes", "Open", date(2024, 4, 1)),
    ]
    for i, (task, status, due) in enumerate(tasks, start=2):
        ws.cell(row=i, column=1, value=task)
        ws.cell(row=i, column=2, value=status)
        dcell = ws.cell(row=i, column=3, value=due)
        dcell.number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=4, value=f"=C{i}-TODAY()")
        ws.cell(row=i, column=5, value=f'=IF(B{i}="Done",TRUE,FALSE)')
    dv = DataValidation(type="list", formula1='"Open,In Progress,Done"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"B2:B{len(tasks) + 1}")
    ws.conditional_formatting.add(
        f"B2:B{len(tasks) + 1}",
        CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    done_row = len(tasks) + 3
    ws.cell(row=done_row, column=1, value="Completed count")
    ws.cell(row=done_row, column=2, value=f'=COUNTIF(B2:B{len(tasks) + 1},"Done")')
    style_header_row(ws, last_col=5)
    wb.save(out("06_todo_checklist.xlsx"))
    record(
        "06_todo_checklist.xlsx", "2-simple", "To-do checklist",
        "Dropdown data validation, conditional-formatting highlight, and two functions outside "
        "the portable profile (TODAY, COUNTIF) alongside a supported IF.",
        ["data validation list", "conditional formatting", "TODAY (unsupported)", "COUNTIF (unsupported)"],
    )


def build_07_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.merge_cells("A1:E1")
    ws["A1"] = "INVOICE"
    ws["A1"].font = Font(bold=True, size=20)
    ws["A3"] = "Bill To:"
    ws["B3"] = "Riverside Design Co."
    ws["A4"] = "Invoice #:"
    ws["B4"] = "INV-1042"
    ws["A5"] = "Date:"
    ws["B5"] = date(2024, 3, 4)
    ws["B5"].number_format = "yyyy-mm-dd"

    ws.append([])
    header_row = 7
    ws.cell(row=header_row, column=1, value="Description")
    ws.cell(row=header_row, column=2, value="Qty")
    ws.cell(row=header_row, column=3, value="Unit Price")
    ws.cell(row=header_row, column=4, value="Line Total")
    style_header_row(ws, row=header_row, last_col=4)

    items = [("Design consultation", 4, 125.0), ("Logo revisions", 3, 80.0), ("Brand guide (PDF)", 1, 450.0)]
    r = header_row + 1
    for desc, qty, price in items:
        ws.cell(row=r, column=1, value=desc)
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=3, value=price).number_format = CURRENCY_USD
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = CURRENCY_USD
        r += 1

    ws.cell(row=r + 1, column=3, value="Subtotal")
    ws.cell(row=r + 1, column=4, value=f"=SUM(D{header_row + 1}:D{r - 1})").number_format = CURRENCY_USD
    ws.cell(row=r + 2, column=3, value="Tax (8%)")
    ws.cell(row=r + 2, column=4, value=f"=D{r + 1}*0.08").number_format = CURRENCY_USD
    ws.cell(row=r + 3, column=3, value="Total").font = Font(bold=True)
    total_cell = ws.cell(row=r + 3, column=4, value=f"=D{r + 1}+D{r + 2}")
    total_cell.font = Font(bold=True)
    total_cell.number_format = CURRENCY_USD

    for row in range(header_row, r):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
    ws.column_dimensions["A"].width = 26
    wb.save(out("07_invoice.xlsx"))
    record(
        "07_invoice.xlsx", "2-simple", "Client invoice",
        "Merged header, a line-item table with per-row multiplication, SUM subtotal, tax "
        "formula, and grand total. Borders and currency formatting throughout.",
        ["merged cells", "multiplication formula", "SUM", "borders"],
    )


def build_08_unit_conversion():
    wb = Workbook()
    ws = wb.active
    ws.title = "Convert"
    ws["A1"] = "km to miles factor"
    ws["B1"] = 0.621371
    wb.defined_names.add(DefinedName("km_to_miles", attr_text="Convert!$B$1"))
    ws.append([])
    ws.cell(row=3, column=1, value="Kilometers")
    ws.cell(row=3, column=2, value="Miles")
    style_header_row(ws, row=3, last_col=2)
    for i in range(10):
        r = 4 + i
        km = (i + 1) * 5
        ws.cell(row=r, column=1, value=km)
        ws.cell(row=r, column=2, value=f"=A{r}*km_to_miles").number_format = "0.000"
    wb.save(out("08_unit_conversion.xlsx"))
    record(
        "08_unit_conversion.xlsx", "2-simple", "Unit conversion table",
        "A single workbook-scoped defined name (km_to_miles) referenced by ten independent "
        "formulas -- tests one-to-many defined-name fan-out.",
        ["defined name", "multiplication formula"],
    )


def build_09_loan_payment_calculator():
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan"
    labels = ["Principal", "Annual Rate", "Term (years)", "Monthly Rate", "Number of Payments", "Monthly Payment"]
    for i, label in enumerate(labels, start=1):
        ws.cell(row=i, column=1, value=label)
    ws["B1"] = 20000
    ws["B1"].number_format = CURRENCY_USD
    ws["B2"] = 0.065
    ws["B2"].number_format = PERCENT
    ws["B3"] = 5
    ws["B4"] = "=B2/12"
    ws["B4"].number_format = "0.0000%"
    ws["B5"] = "=B3*12"
    ws["B6"] = "=PMT(B4,B5,-B1)"
    ws["B6"].number_format = CURRENCY_USD
    ws.column_dimensions["A"].width = 20
    wb.save(out("09_loan_payment_calculator.xlsx"))
    record(
        "09_loan_payment_calculator.xlsx", "2-simple", "Loan payment calculator",
        "Single-sheet amortization inputs feeding a PMT() financial formula -- PMT is outside "
        "the portable-a1@1 profile, so this exercises the lossy/replaced formula path.",
        ["PMT (unsupported)", "percent format", "chained scalar formulas"],
    )


# ==========================================================================
# TIER 3 -- Intermediate
# ==========================================================================

def build_10_multi_sheet_sales_report():
    wb = Workbook()
    raw = wb.active
    raw.title = "RawData"
    raw.append(["Region", "Rep", "Product", "Amount"])
    import random
    random.seed(11)
    regions = ["North", "South", "West"]
    reps = {"North": ["Ana", "Bo"], "South": ["Cy", "Dee"], "West": ["Emi", "Fox"]}
    products = ["Widget", "Gadget", "Gizmo"]
    r = 2
    for _ in range(45):
        region = random.choice(regions)
        raw.append([region, random.choice(reps[region]), random.choice(products), round(random.uniform(50, 5000), 2)])
        r += 1
    last_raw = r - 1
    for col, w in zip("ABCD", (10, 8, 10, 12)):
        raw.column_dimensions[col].width = w

    for region in regions:
        ws = wb.create_sheet(region)
        ws["A1"] = f"{region} region summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "Total sales"
        ws["B3"] = f'=SUMIF(RawData!A2:A{last_raw},"{region}",RawData!D2:D{last_raw})'
        ws["B3"].number_format = CURRENCY_USD
        ws["A4"] = "Transaction count"
        ws["B4"] = f'=COUNTIF(RawData!A2:A{last_raw},"{region}")'
        ws["A5"] = "Average sale"
        ws["B5"] = f'=AVERAGEIF(RawData!A2:A{last_raw},"{region}",RawData!D2:D{last_raw})'

    summary = wb.create_sheet("Summary")
    summary.append(["Region", "Total"])
    for i, region in enumerate(regions, start=2):
        summary.cell(row=i, column=1, value=region)
        summary.cell(row=i, column=2, value=f"={region}!B3")
        summary.cell(row=i, column=2).number_format = CURRENCY_USD
    summary.cell(row=len(regions) + 2, column=1, value="Grand total").font = Font(bold=True)
    summary.cell(row=len(regions) + 2, column=2, value=f"=SUM(B2:B{len(regions) + 1})").number_format = CURRENCY_USD
    style_header_row(summary, last_col=2)
    style_header_row(raw, last_col=4)

    wb.save(out("10_multi_sheet_sales_report.xlsx"))
    record(
        "10_multi_sheet_sales_report.xlsx", "3-intermediate", "Multi-sheet sales report",
        "Five sheets (RawData + 3 region sheets + Summary) chained by cross-sheet references, "
        "SUMIF/COUNTIF/AVERAGEIF per region, then a Summary sheet that reads each region "
        "sheet's own formula result.",
        ["5 sheets", "cross-sheet refs", "SUMIF/COUNTIF/AVERAGEIF (unsupported)", "SUM"],
    )


def build_11_lookup_directory():
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["ID", "Name", "Dept", "Manager", "Salary"])
    people = [
        (101, "Priya Nair", "Engineering", "Sam Ortiz", 118000),
        (102, "Sam Ortiz", "Engineering", "—", 165000),
        (103, "Lucia Fernandez", "Design", "Noah Kim", 104000),
        (104, "Noah Kim", "Design", "—", 142000),
        (105, "Owen Bright", "Sales", "Maya Chen", 96000),
        (106, "Maya Chen", "Sales", "—", 150000),
        (107, "Ravi Deshpande", "Engineering", "Sam Ortiz", 121000),
        (108, "Ingrid Solberg", "Marketing", "Maya Chen", 99000),
        (109, "Tomas Varga", "Engineering", "Sam Ortiz", 108500),
        (110, "Aiko Tanaka", "Design", "Noah Kim", 111000),
    ]
    for row in people:
        emp.append(row)
    emp["E2"].number_format = CURRENCY_USD
    last = len(people) + 1
    for row in range(2, last + 1):
        emp.cell(row=row, column=5).number_format = CURRENCY_USD
    style_header_row(emp, last_col=5)

    lk = wb.create_sheet("Lookups")
    lk.append(["Lookup ID", "VLOOKUP Name", "HLOOKUP-style", "INDEX/MATCH Salary", "XLOOKUP Dept"])
    ids = [103, 106, 109]
    for i, emp_id in enumerate(ids, start=2):
        lk.cell(row=i, column=1, value=emp_id)
        lk.cell(row=i, column=2, value=f"=VLOOKUP(A{i},Employees!A:E,2,FALSE)")
        lk.cell(row=i, column=3, value=f"=HLOOKUP(A{i},Employees!A1:E{last},2,FALSE)")
        lk.cell(
            row=i, column=4,
            value=f"=INDEX(Employees!E:E,MATCH(A{i},Employees!A:A,0))",
        ).number_format = CURRENCY_USD
        lk.cell(row=i, column=5, value=f"=XLOOKUP(A{i},Employees!A:A,Employees!C:C)")
    style_header_row(lk, last_col=5)
    for col in "ABCDE":
        lk.column_dimensions[col].width = 18

    wb.save(out("11_lookup_directory.xlsx"))
    record(
        "11_lookup_directory.xlsx", "3-intermediate", "Employee lookup directory",
        "Two sheets: an Employees table and a Lookups sheet exercising VLOOKUP, HLOOKUP, "
        "INDEX+MATCH (both in the portable profile), and XLOOKUP (not) against it.",
        ["VLOOKUP (unsupported)", "HLOOKUP (unsupported)", "INDEX", "MATCH", "XLOOKUP (unsupported)"],
    )


def build_12_named_ranges_tax_calc():
    wb = Workbook()
    assumptions = wb.active
    assumptions.title = "Assumptions"
    assumptions["A1"] = "Tax rate"
    assumptions["B1"] = 0.22
    assumptions["B1"].number_format = PERCENT
    assumptions["A2"] = "Standard deduction"
    assumptions["B2"] = 13850
    assumptions["B2"].number_format = CURRENCY_USD
    wb.defined_names.add(DefinedName("tax_rate", attr_text="Assumptions!$B$1"))
    wb.defined_names.add(DefinedName("standard_deduction", attr_text="Assumptions!$B$2"))
    # Sheet-scoped (local) defined names are exercised separately in
    # 32_sheet_scoped_defined_name.xlsx -- crates/marksheet-convert currently
    # rejects the *whole file* when one is present ("sheet-local defined
    # names are outside the initial profile"), so keeping one here would
    # prevent this file from demonstrating ordinary named-range formulas.

    calc = wb.create_sheet("Calc")
    calc.append(["Employee", "Gross Income", "Taxable Income", "Tax Owed", "Net Income"])
    people = [("Priya", 118000), ("Owen", 96000), ("Ingrid", 99000)]
    for i, (name, gross) in enumerate(people, start=2):
        calc.cell(row=i, column=1, value=name)
        calc.cell(row=i, column=2, value=gross).number_format = CURRENCY_USD
        calc.cell(row=i, column=3, value=f"=B{i}-standard_deduction").number_format = CURRENCY_USD
        calc.cell(row=i, column=4, value=f"=C{i}*tax_rate").number_format = CURRENCY_USD
        calc.cell(row=i, column=5, value=f"=B{i}-D{i}").number_format = CURRENCY_USD
    style_header_row(calc, last_col=5)

    wb.save(out("12_named_ranges_tax_calc.xlsx"))
    record(
        "12_named_ranges_tax_calc.xlsx", "3-intermediate", "Named-range tax calculator",
        "Two workbook-scoped defined names (tax_rate, standard_deduction) resolved inside "
        "ordinary arithmetic formulas. (Sheet-scoped names are exercised separately in "
        "32_sheet_scoped_defined_name.xlsx.)",
        ["workbook-scoped defined names", "arithmetic formulas"],
    )


def build_13_excel_table_orders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["OrderID", "Customer", "Qty", "Price", "Total"])
    rows = [
        (1001, "Acme Corp", 12, 9.5), (1002, "Bright Foods", 4, 42.0), (1003, "Cedar Studio", 20, 3.25),
        (1004, "Acme Corp", 6, 9.5), (1005, "Dune Labs", 1, 899.0), (1006, "Bright Foods", 15, 4.75),
        (1007, "Cedar Studio", 8, 12.0), (1008, "Emberly LLC", 3, 150.0),
    ]
    r = 2
    for order_id, customer, qty, price in rows:
        ws.cell(row=r, column=1, value=order_id)
        ws.cell(row=r, column=2, value=customer)
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value=price).number_format = CURRENCY_USD
        ws.cell(row=r, column=5, value="=[@Qty]*[@Price]").number_format = CURRENCY_USD
        r += 1
    last = r - 1
    table = Table(displayName="Orders", ref=f"A1:E{last}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False,
    )
    ws.add_table(table)
    for col, w in zip("ABCDE", (10, 16, 6, 10, 12)):
        ws.column_dimensions[col].width = w
    wb.save(out("13_excel_table_orders.xlsx"))
    record(
        "13_excel_table_orders.xlsx", "3-intermediate", "Excel Table with calculated column",
        "A real Excel Table (ListObject) named 'Orders' with a calculated column driven by "
        "structured references ([@Qty]*[@Price]).",
        ["Excel Table", "structured references", "calculated column"],
    )


def build_14_conditional_formatting_dashboard():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Metric", "Value", "Target", "% of Target"])
    metrics = [
        ("New signups", 842, 900), ("Churn (lower is better)", 34, 50), ("NPS", 61, 70),
        ("Support tickets closed", 210, 200), ("Uptime %", 99.95, 99.9), ("Revenue ($k)", 412, 450),
    ]
    for i, (name, value, target) in enumerate(metrics, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=3, value=target)
        ws.cell(row=i, column=4, value=f"=B{i}/C{i}").number_format = PERCENT
    last = len(metrics) + 1
    style_header_row(ws, last_col=4)

    ws.conditional_formatting.add(
        f"B2:B{last}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )
    ws.conditional_formatting.add(
        f"D2:D{last}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.2, color="638EC6"),
    )
    ws.conditional_formatting.add(
        f"D2:D{last}",
        IconSetRule(icon_style="3TrafficLights1", type="percent", values=[0, 33, 67]),
    )
    ws.conditional_formatting.add(
        f"D2:D{last}",
        CellIsRule(operator="lessThan", formula=["1"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    wb.save(out("14_conditional_formatting_dashboard.xlsx"))
    record(
        "14_conditional_formatting_dashboard.xlsx", "3-intermediate", "KPI dashboard with conditional formatting",
        "Color scale, data bar, icon set, and a formula-based cell rule stacked on the same "
        "KPI table, alongside a plain division formula.",
        ["color scale", "data bar", "icon set", "cell-value rule", "division formula"],
    )


def build_15_data_validation_form():
    wb = Workbook()
    ws = wb.active
    ws.title = "Form"
    labels = [
        "Status (list)", "Priority 1-5 (whole number)", "Discount 0-1 (decimal)",
        "Start date (date range)", "Notes (text length <= 20)", "Confirm > 100 (custom formula)",
    ]
    for i, label in enumerate(labels, start=1):
        ws.cell(row=i, column=1, value=label)

    dv_list = DataValidation(type="list", formula1='"Draft,Submitted,Approved,Rejected"', allow_blank=True)
    dv_list.prompt = "Choose one"
    dv_list.promptTitle = "Status"
    ws.add_data_validation(dv_list)
    dv_list.add("B1")

    dv_int = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
    dv_int.error = "Enter 1-5"
    dv_int.errorTitle = "Invalid priority"
    ws.add_data_validation(dv_int)
    dv_int.add("B2")

    dv_dec = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    ws.add_data_validation(dv_dec)
    dv_dec.add("B3")

    dv_date = DataValidation(type="date", operator="greaterThanOrEqual", formula1=date(2024, 1, 1))
    ws.add_data_validation(dv_date)
    dv_date.add("B4")
    ws["B4"].number_format = "yyyy-mm-dd"

    dv_len = DataValidation(type="textLength", operator="lessThanOrEqual", formula1=20)
    ws.add_data_validation(dv_len)
    dv_len.add("B5")

    dv_custom = DataValidation(type="custom", formula1="B6>100")
    ws.add_data_validation(dv_custom)
    dv_custom.add("B6")

    ws.column_dimensions["A"].width = 34
    wb.save(out("15_data_validation_form.xlsx"))
    record(
        "15_data_validation_form.xlsx", "3-intermediate", "Data validation form",
        "One cell each for list, whole-number range, decimal range, date range, text-length, "
        "and custom-formula data validation rules.",
        ["list validation", "whole/decimal range validation", "date validation", "custom formula validation"],
    )


def build_16_frozen_panes_large_grid():
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid"
    n_rows, n_cols = 200, 14
    ws.cell(row=1, column=1, value="Row")
    for c in range(2, n_cols + 2):
        ws.cell(row=1, column=c, value=f"Metric {c - 1}")
    import random
    random.seed(3)
    for r in range(2, n_rows + 2):
        ws.cell(row=r, column=1, value=f"Item {r - 1}")
        for c in range(2, n_cols + 2):
            ws.cell(row=r, column=c, value=round(random.uniform(-100, 100), 2))
    total_row = n_rows + 3
    ws.cell(row=total_row, column=1, value="Column totals").font = Font(bold=True)
    for c in range(2, n_cols + 2):
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}2:{col_letter}{n_rows + 1})")
    ws.freeze_panes = "B2"
    style_header_row(ws, last_col=n_cols + 1)
    ws.column_dimensions["A"].width = 14
    for c in range(2, n_cols + 2):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 11
    ws.conditional_formatting.add(
        f"B2:{ws.cell(row=1, column=n_cols + 1).column_letter}{n_rows + 1}",
        ColorScaleRule(start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B"),
    )
    wb.save(out("16_frozen_panes_large_grid.xlsx"))
    record(
        "16_frozen_panes_large_grid.xlsx", "3-intermediate", "Frozen-pane large grid",
        f"{n_rows}x{n_cols} numeric grid with a frozen header row/column, explicit column "
        "widths, a color-scale heat map, and SUM column totals.",
        ["freeze panes", "column widths", "color scale", "SUM", f"{n_rows} rows"],
    )


# ==========================================================================
# TIER 4 -- Complex
# ==========================================================================

def build_17_financial_model_3_statement():
    wb = Workbook()
    a = wb.active
    a.title = "Assumptions"
    a.append(["Assumption", "Value"])
    a.append(["Revenue growth", 0.12])
    a.append(["COGS % of revenue", 0.42])
    a.append(["OpEx % of revenue", 0.28])
    a.append(["D&A % of revenue", 0.05])
    a.append(["Interest expense", 45000])
    a.append(["Tax rate", 0.24])
    for row in range(2, 8):
        if row in (2, 3, 4, 5, 7):
            a.cell(row=row, column=2).number_format = PERCENT
        else:
            a.cell(row=row, column=2).number_format = CURRENCY_USD
    wb.defined_names.add(DefinedName("revenue_growth", attr_text="Assumptions!$B$2"))
    wb.defined_names.add(DefinedName("cogs_pct", attr_text="Assumptions!$B$3"))
    wb.defined_names.add(DefinedName("opex_pct", attr_text="Assumptions!$B$4"))
    wb.defined_names.add(DefinedName("da_pct", attr_text="Assumptions!$B$5"))
    wb.defined_names.add(DefinedName("interest_expense", attr_text="Assumptions!$B$6"))
    wb.defined_names.add(DefinedName("tax_rate", attr_text="Assumptions!$B$7"))

    inc = wb.create_sheet("Income Statement")
    years = ["Year 1", "Year 2", "Year 3"]
    inc.append(["Line"] + years)
    rows = ["Revenue", "COGS", "Gross Profit", "OpEx", "EBITDA", "D&A", "EBIT", "Interest", "Pretax Income", "Tax", "Net Income"]
    for r, label in enumerate(rows, start=2):
        inc.cell(row=r, column=1, value=label)
    inc["B2"] = 2000000
    inc["B2"].number_format = CURRENCY_USD
    for col_idx in range(3, 5):
        col = inc.cell(row=2, column=col_idx).column_letter
        prev = inc.cell(row=2, column=col_idx - 1).column_letter
        inc.cell(row=2, column=col_idx, value=f"={prev}2*(1+revenue_growth)").number_format = CURRENCY_USD
    for col_idx in range(2, 5):
        col = inc.cell(row=1, column=col_idx).column_letter
        inc.cell(row=3, column=col_idx, value=f"={col}2*cogs_pct").number_format = CURRENCY_USD
        inc.cell(row=4, column=col_idx, value=f"={col}2-{col}3").number_format = CURRENCY_USD
        inc.cell(row=5, column=col_idx, value=f"={col}2*opex_pct").number_format = CURRENCY_USD
        inc.cell(row=6, column=col_idx, value=f"={col}4-{col}5").number_format = CURRENCY_USD
        inc.cell(row=7, column=col_idx, value=f"={col}2*da_pct").number_format = CURRENCY_USD
        inc.cell(row=8, column=col_idx, value=f"={col}6-{col}7").number_format = CURRENCY_USD
        inc.cell(row=9, column=col_idx, value="=interest_expense").number_format = CURRENCY_USD
        inc.cell(row=10, column=col_idx, value=f"={col}8-{col}9").number_format = CURRENCY_USD
        inc.cell(row=11, column=col_idx, value=f"={col}10*tax_rate").number_format = CURRENCY_USD
        inc.cell(row=12, column=col_idx, value=f"={col}10-{col}11").number_format = CURRENCY_USD
    style_header_row(inc, last_col=4)

    bs = wb.create_sheet("Balance Sheet")
    bs.append(["Line"] + years)
    bs.append(["Cash", 500000, "=B2+'Cash Flow'!B12", "=C2+'Cash Flow'!C12"])
    bs.append(["Retained Earnings", 300000, "=B3+'Income Statement'!C12", "=C3+'Income Statement'!D12"])
    for row in (2, 3):
        for col in "BCD":
            bs[f"{col}{row}"].number_format = CURRENCY_USD
    style_header_row(bs, last_col=4)

    cf = wb.create_sheet("Cash Flow")
    cf.append(["Line"] + years)
    cf.append(["Net Income", "='Income Statement'!B12", "='Income Statement'!C12", "='Income Statement'!D12"])
    cf.append(["Add back D&A", "='Income Statement'!B7", "='Income Statement'!C7", "='Income Statement'!D7"])
    cf.append(["Free Cash Flow", "=B2+B3", "=C2+C3", "=D2+D3"])
    for row in (2, 3, 4):
        for col in "BCD":
            cf[f"{col}{row}"].number_format = CURRENCY_USD
    cf["A6"] = "NPV @ 10%"
    cf["B6"] = "=NPV(0.1,B4:D4)"
    cf["B6"].number_format = CURRENCY_USD
    cf["A7"] = "IRR"
    cf["B7"] = "=IRR(B4:D4)"
    cf["B7"].number_format = PERCENT
    style_header_row(cf, last_col=4)

    wb.save(out("17_financial_model_3_statement.xlsx"))
    record(
        "17_financial_model_3_statement.xlsx", "4-complex", "Three-statement financial model",
        "Assumptions + Income Statement + Balance Sheet + Cash Flow, chained with defined-name "
        "driven formulas across sheets, plus NPV and IRR.",
        ["4 sheets", "defined names", "cross-sheet chains", "NPV (unsupported)", "IRR (unsupported)"],
    )


def build_18_large_dataset():
    wb = Workbook()
    tx = wb.active
    tx.title = "Transactions"
    tx.append(["Date", "Region", "Category", "Amount"])
    import random
    random.seed(42)
    regions = ["North", "South", "East", "West"]
    categories = ["Hardware", "Software", "Services", "Support"]
    n = 2000
    start = date(2023, 1, 1)
    for i in range(n):
        d = start + timedelta(days=random.randint(0, 700))
        tx.append([d, random.choice(regions), random.choice(categories), round(random.uniform(20, 8000), 2)])
    for r in range(2, n + 2):
        tx.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        tx.cell(row=r, column=4).number_format = CURRENCY_USD
    style_header_row(tx, last_col=4)
    last = n + 1

    summary = wb.create_sheet("Summary")
    summary.append(["Region", "Category", "Total", "Count", "Average"])
    r = 2
    for region in regions:
        for cat in categories:
            summary.cell(row=r, column=1, value=region)
            summary.cell(row=r, column=2, value=cat)
            summary.cell(
                row=r, column=3,
                value=(
                    f'=SUMIFS(Transactions!D2:D{last},Transactions!B2:B{last},A{r},'
                    f"Transactions!C2:C{last},B{r})"
                ),
            ).number_format = CURRENCY_USD
            summary.cell(
                row=r, column=4,
                value=f'=COUNTIFS(Transactions!B2:B{last},A{r},Transactions!C2:C{last},B{r})',
            )
            summary.cell(
                row=r, column=5,
                value=(
                    f'=AVERAGEIFS(Transactions!D2:D{last},Transactions!B2:B{last},A{r},'
                    f"Transactions!C2:C{last},B{r})"
                ),
            ).number_format = CURRENCY_USD
            r += 1
    style_header_row(summary, last_col=5)
    wb.save(out("18_large_dataset_2000_rows.xlsx"))
    record(
        "18_large_dataset_2000_rows.xlsx", "4-complex", "Large transaction dataset",
        f"{n} rows of transactional data plus a 16-row SUMIFS/COUNTIFS/AVERAGEIFS cross-tab "
        "summary reading the full range -- a scale/performance stress case.",
        [f"{n} rows", "SUMIFS/COUNTIFS/AVERAGEIFS (unsupported)", "cross-sheet range refs"],
    )


def build_19_array_and_dynamic_formulas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Arrays"
    write_seed_block(ws)
    rows = [
        ("SUMPRODUCT (legacy array-capable)", "=SUMPRODUCT(H2:H11,I2:I11)"),
        ("Array-entered conditional sum", "=SUM(IF(H2:H11>0,H2:H11))"),
        ("FILTER (dynamic array)", "=FILTER(J2:J11,L2:L11)"),
        ("SORT (dynamic array)", "=SORT(H2:H11,1,-1)"),
        ("SORTBY (dynamic array)", "=SORTBY(J2:J11,H2:H11)"),
        ("UNIQUE (dynamic array)", "=UNIQUE(J2:J11)"),
        ("SEQUENCE (dynamic array)", "=SEQUENCE(5,1,1,1)"),
        ("Spilled range reference", "=H2#"),
    ]
    for i, (label, formula) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)
    style_header_row(ws, last_col=2)
    wb.save(out("19_array_and_dynamic_formulas.xlsx"))
    record(
        "19_array_and_dynamic_formulas.xlsx", "4-complex", "Array and dynamic-array formulas",
        "Legacy CSE-style array formulas (SUMPRODUCT, array-entered SUM/IF) next to modern "
        "dynamic-array functions (FILTER, SORT, SORTBY, UNIQUE, SEQUENCE) and a spill reference.",
        ["SUMPRODUCT (unsupported)", "array formula", "FILTER/SORT/UNIQUE/SEQUENCE (unsupported)", "spill reference"],
    )


def build_20_error_handling_showcase():
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    ws.append(["Case", "Formula", "Wrapped in IFERROR/IFNA/ISERROR"])
    cases = [
        ("#DIV/0!", "=1/0", "=IFERROR(1/0,\"caught\")"),
        ("#N/A", '=MATCH("missing",{"a","b","c"},0)', '=IFNA(MATCH("missing",{"a","b","c"},0),"caught")'),
        ("#VALUE!", '="text"+1', '=IFERROR("text"+1,"caught")'),
        ("#REF!", "=#REF!", "=IFERROR(#REF!,\"caught\")"),
        ("#NAME?", "=undefined_name_xyz*2", '=IFERROR(undefined_name_xyz*2,"caught")'),
        ("#NUM!", "=SQRT(-1)", '=IFERROR(SQRT(-1),"caught")'),
        # #NULL! is kept out of this list because SPEC.md section 13's
        # required runtime-failure table omits it (only DIV/0!, N/A, NAME?,
        # NUM!, REF!, VALUE!, and CIRC! are required) -- see
        # 33_null_error_token.xlsx for that case on its own.
        ("ISERROR check", "=ISERROR(1/0)", None),
        ("ISNA check", '=ISNA(NA())', None),
    ]
    for i, (label, formula, wrapped) in enumerate(cases, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)
        if wrapped:
            ws.cell(row=i, column=3, value=wrapped)
    style_header_row(ws, last_col=3)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    wb.save(out("20_error_handling_showcase.xlsx"))
    record(
        "20_error_handling_showcase.xlsx", "4-complex", "Every Excel runtime error, deliberately triggered",
        "One row per error class SPEC.md section 13 actually requires (#DIV/0!, #N/A, "
        "#VALUE!, #REF!, #NAME?, #NUM!) plus the matching IFERROR/IFNA-wrapped version and "
        "ISERROR/ISNA checks. (#NULL! is out of scope for portable-a1@1 and is exercised "
        "separately in 33_null_error_token.xlsx.)",
        ["#DIV/0!", "#N/A", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "IFERROR", "IFNA", "ISERROR"],
    )


def build_21_circular_reference():
    wb = Workbook()
    ws = wb.active
    ws.title = "Circular"
    ws["A1"] = "Mutual circular pair"
    ws["A2"] = "=B2+1"
    ws["B2"] = "=A2+1"
    ws["A4"] = "Self-loop"
    ws["A5"] = "=A5+1"
    ws["A7"] = "Downstream of the circular pair (should also surface #CIRC!)"
    ws["A8"] = "=A2*2"
    wb.save(out("21_circular_reference.xlsx"))
    record(
        "21_circular_reference.xlsx", "4-complex", "Circular references",
        "A mutual two-cell circular pair, a single self-loop, and a downstream formula that "
        "depends on the circular pair -- SPEC.md requires every cell in the strongly connected "
        "component (and its dependents) to resolve to #CIRC!.",
        ["mutual circular reference", "self-loop", "#CIRC! propagation"],
    )


def build_22_charts():
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart"
    ws.append(["Month", "Revenue", "Expenses", "Profit"])
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    import random
    random.seed(5)
    for i, m in enumerate(months, start=2):
        revenue = random.randint(8000, 15000)
        expenses = random.randint(5000, 9000)
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=revenue)
        ws.cell(row=i, column=3, value=expenses)
        ws.cell(row=i, column=4, value=f"=B{i}-C{i}")
    last = len(months) + 1
    style_header_row(ws, last_col=4)

    bar = BarChart()
    bar.title = "Revenue vs Expenses"
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "F2")

    line = LineChart()
    line.title = "Profit trend"
    data2 = Reference(ws, min_col=4, min_row=1, max_row=last)
    line.add_data(data2, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "F18")

    pie = PieChart()
    pie.title = "June split"
    pie_data = Reference(ws, min_col=2, max_col=3, min_row=last, max_row=last)
    pie_cats = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1)
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_cats)
    ws.add_chart(pie, "F34")

    wb.save(out("22_charts.xlsx"))
    record(
        "22_charts.xlsx", "4-complex", "Embedded charts",
        "A bar chart, line chart, and pie chart embedded alongside their source data -- "
        "exercises the Chart/Unsupported import path (crates/marksheet-convert import.rs "
        "treats any xl/charts/* part as outside the initial import profile).",
        ["bar chart", "line chart", "pie chart"],
    )


def build_23_pivot_table_source():
    wb = Workbook()
    tx = wb.active
    tx.title = "Transactions"
    tx.append(["Region", "Amount", "Date", "Category"])
    import random
    random.seed(9)
    regions = ["North", "South", "East", "West", "North", "South", "East", "West", "North", "South"]
    for i, region in enumerate(regions, start=2):
        tx.cell(row=i, column=1, value=region)
        tx.cell(row=i, column=2, value=round(random.uniform(100, 5000), 2))
        tx.cell(row=i, column=3, value=date(2024, 1, ((i - 2) % 28) + 1))
        tx.cell(row=i, column=4, value=random.choice(["Hardware", "Services"]))
    style_header_row(tx, last_col=4)
    path = out("23_pivot_table_source.xlsx")
    wb.save(path)
    add_pivot_table_stub(path)
    record(
        "23_pivot_table_source.xlsx", "4-complex", "Pivot table (stub) over source data",
        "Real source data plus a minimal, hand-authored xl/pivotCache + xl/pivotTables part "
        "pair -- enough to exercise Marksheet's PivotTable/Unsupported detection, though it is "
        "not a substitute for a real Excel-authored pivot table (see README.md).",
        ["pivot table (synthetic stub)", "source data table"],
    )


# ==========================================================================
# TIER 5 -- Edge cases / stress / interop
# ==========================================================================

def build_24_unicode_and_special_characters():
    wb = Workbook()
    ws = wb.active
    ws.title = "Données 数据"
    ws.append(["Language", "Text", "Notes"])
    samples = [
        ("Emoji", "Q1 report ✅ 🚀📈", ""),
        ("Chinese", "销售报告", "Simplified Chinese"),
        ("Japanese", "売上報告書", "Japanese"),
        ("Korean", "매출 보고서", "Korean"),
        ("Arabic (RTL)", "تقرير المبيعات", "Right-to-left"),
        ("Hebrew (RTL)", "דוח מכירות", "Right-to-left"),
        ("Accented Latin", "Café Müller — naïve résumé", "Combining/precomposed accents"),
        ("Escaped quotes", 'She said ""hi"" to me', "Literal via formula, see B9"),
    ]
    for i, (lang, text, note) in enumerate(samples, start=2):
        ws.cell(row=i, column=1, value=lang)
        if lang == "Escaped quotes":
            ws.cell(row=i, column=2, value='=""She said ""hi"" to me""')
        else:
            ws.cell(row=i, column=2, value=text)
        ws.cell(row=i, column=3, value=note)
    style_header_row(ws, last_col=3)

    ws2 = wb.create_sheet("Ref Target")
    ws2["A1"] = "value on a unicode-named sheet"
    ws2["B1"] = 99
    ws["A11"] = "Cross-sheet ref to unicode sheet name"
    ws["B11"] = "='Données 数据'!B11"
    ws["A12"] = "Reference into 'Ref Target'"
    ws["B12"] = "='Ref Target'!B1*2"
    wb.save(out("24_unicode_and_special_characters.xlsx"))
    record(
        "24_unicode_and_special_characters.xlsx", "5-edge-case", "Unicode and special characters",
        "Emoji, CJK, RTL Arabic/Hebrew, accented Latin, a doubled-double-quote escape sequence, "
        "and a unicode sheet name referenced from a formula.",
        ["unicode text", "RTL text", "unicode sheet name", "escaped quotes in formula"],
    )


def build_25_many_sheets():
    wb = Workbook()
    first = wb.active
    first.title = "Sheet01"
    first["A1"] = 1
    n = 50
    for i in range(2, n + 1):
        ws = wb.create_sheet(f"Sheet{i:02d}")
        ws["A1"] = f"=Sheet{i - 1:02d}!A1+1"
    index = wb.create_sheet("Index", 0)
    index["A1"] = "Chained total (should equal 50)"
    index["B1"] = f"=Sheet{n:02d}!A1"
    wb.save(out("25_many_sheets.xlsx"))
    record(
        "25_many_sheets.xlsx", "5-edge-case", "Many sheets",
        f"{n + 1} sheets total, each formula-chained to the previous one (Sheet02!A1 = "
        "Sheet01!A1+1, and so on), read back from an Index sheet.",
        [f"{n + 1} sheets", "sheet-to-sheet formula chain"],
    )


def build_26_merged_cells_comments_hyperlinks():
    wb = Workbook()
    ws = wb.active
    ws.title = "Doc"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Quarterly Report"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:B2")
    ws["A2"] = "Prepared by"
    ws.merge_cells("C2:D2")
    ws["C2"] = "Finance Team"
    ws.merge_cells("A4:B5")
    ws["A4"] = "2x2 merged block"

    ws["A7"] = "Commented cell"
    ws["A7"].comment = Comment("This figure is provisional pending audit.", "Finance Bot")
    ws["B7"] = 41250
    ws["B7"].comment = Comment("Confirm with controller before publishing.", "Finance Bot")

    # An *external* hyperlink (TargetMode="External") is deliberately not
    # included here: crates/marksheet-convert/src/xlsx/package.rs rejects
    # any package relationship with TargetMode="External" outright ("external
    # OOXML relationships are rejected"), before per-feature handling even
    # runs -- so it takes down the whole file rather than just the
    # hyperlink. See 34_external_hyperlink.xlsx for that isolated. Note this
    # applies even to a same-sheet link if authored via the plain-string
    # `.hyperlink = "#Doc!A1"` API -- openpyxl always emits that as a
    # TargetMode="External" relationship regardless of the "#" prefix. The
    # `location=` form below instead emits a relationship-free
    # <hyperlink location="..."/> element, which imports cleanly.
    ws["A9"] = "(external hyperlink intentionally omitted -- see README.md)"
    ws["A10"] = "Internal link"
    ws["A10"].hyperlink = Hyperlink(ref="A10", location="Doc!A1", display="Internal link")
    ws["A10"].font = Font(color="0563C1", underline="single")

    ws.column_dimensions["C"].hidden = True
    ws.row_dimensions[3].height = 5
    wb.save(out("26_merged_cells_comments_hyperlinks.xlsx"))
    record(
        "26_merged_cells_comments_hyperlinks.xlsx", "5-edge-case", "Merged cells, comments, and hyperlinks",
        "A merged title band, a 2x2 merged block, two cell comments, an internal same-sheet "
        "hyperlink, and a hidden column. (An external hyperlink is exercised separately in "
        "34_external_hyperlink.xlsx -- see its notes.)",
        ["merged cells", "cell comments", "internal hyperlink", "hidden column"],
    )


def build_27_macro_enabled():
    wb = Workbook()
    ws = wb.active
    ws.title = "Macro Demo"
    ws["A1"] = "This workbook is flagged macro-enabled at the package level."
    ws["A2"] = "See README.md: the VBA project bytes are a placeholder, not real bytecode."
    ws["A4"] = "Data a macro might have touched"
    ws["A5"] = "Total"
    ws["B5"] = "=SUM(C5:E5)"
    ws["C5"], ws["D5"], ws["E5"] = 10, 20, 30
    path = out("27_macro_enabled.xlsx")
    wb.save(path)
    xlsm_path = out("27_macro_enabled.xlsm")
    make_macro_enabled(path, xlsm_path)
    record(
        "27_macro_enabled.xlsm", "5-edge-case", "Macro-enabled workbook",
        "Package-level content type flipped to macroEnabled.main plus a placeholder "
        "xl/vbaProject.bin part -- exercises is_macro_enabled() and the vbaproject-name "
        "detection path in crates/marksheet-convert/src/xlsx/import.rs.",
        ["macro-enabled content type", "vbaProject.bin stub"],
    )


def build_28_number_formats_and_styles():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formats"
    ws.append(["Label", "Value", "Format code applied"])
    formats = [
        ("US currency", 1234.5, CURRENCY_USD),
        ("EUR currency", 1234.5, '#,##0.00" €"'),
        ("GBP currency", 1234.5, '"£"#,##0.00'),
        ("Accounting (paren negatives)", -1234.5, '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'),
        ("Percentage", 0.4567, "0.00%"),
        ("Scientific", 123456789, "0.00E+00"),
        ("Fraction", 1.25, "# ?/?"),
        ("Thousands, no decimals", 1234567, "#,##0"),
        # A literal "[Red]" custom-color directive is deliberately avoided
        # here: crates/marksheet-convert/src/xlsx/import.rs's custom-format
        # sniffer (apply_number_format) flags any format code containing the
        # letter "y" or "d" as a date/time format, so a format as ordinary as
        # "0.00;[Red]-0.00" gets misread as a date purely because "Red"
        # contains a "d" -- and a negative value under a date-classified
        # format is then rejected outright. See 35_custom_format_date_false_positive.xlsx.
        ("Custom negative in parens", -42.5, "0.00;(0.00)"),
        ("Date (ISO)", date(2024, 7, 4), "yyyy-mm-dd"),
        ("Date (US)", date(2024, 7, 4), "mm/dd/yyyy"),
        ("Date (long)", date(2024, 7, 4), "dddd, mmmm d, yyyy"),
        ("Time", datetime(2024, 7, 4, 15, 30), "hh:mm AM/PM"),
        ("Date + time", datetime(2024, 7, 4, 15, 30), "yyyy-mm-dd hh:mm"),
        ("Custom text wrapper", "north", '"Region: "@'),
    ]
    for i, (label, value, fmt) in enumerate(formats, start=2):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=value)
        c.number_format = fmt
        ws.cell(row=i, column=3, value=fmt)
    style_header_row(ws, last_col=3)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 40

    styled = wb.create_sheet("Styles")
    styled["A1"] = "Bold red on yellow"
    styled["A1"].font = Font(bold=True, color="FF0000")
    styled["A1"].fill = PatternFill("solid", fgColor="FFFF00")
    styled["A2"] = "Italic underline"
    styled["A2"].font = Font(italic=True, underline="single")
    styled["A3"] = "Thin border box"
    styled["A3"].border = THIN_BORDER
    styled["A4"] = "Double border"
    dbl = Side(style="double")
    styled["A4"].border = Border(top=dbl, bottom=dbl, left=dbl, right=dbl)
    for r, level in [(6, 1), (7, 2), (8, 2), (9, 1)]:
        styled.cell(row=r, column=1, value=f"Outline level {level} row")
        styled.row_dimensions[r].outlineLevel = level
    styled.column_dimensions["B"].outlineLevel = 1
    styled.cell(row=1, column=2, value="outlined column")

    wb.save(out("28_number_formats_and_styles.xlsx"))
    record(
        "28_number_formats_and_styles.xlsx", "5-edge-case", "Number formats and styles matrix",
        "Currency (3 locales), accounting, percentage, scientific, fraction, a custom "
        "parenthesized-negative format, four date/time formats, and a custom text wrapper, "
        "plus a second sheet of font/fill/border styling and row/column outline (grouping) "
        "levels. (A '[Red]'-colored negative format is exercised separately in "
        "35_custom_format_date_false_positive.xlsx -- see its notes.)",
        ["currency formats", "accounting format", "scientific/fraction formats", "custom formats", "borders", "outline grouping"],
    )


def build_29_external_links_and_name_errors():
    wb = Workbook()
    ws = wb.active
    ws.title = "External"
    ws["A1"] = "Formula referencing another workbook (authored, not a live link)"
    ws["A2"] = "='[Budget-2023.xlsx]Summary'!$B$2"
    ws["A4"] = "Formula referencing an undefined name"
    ws["A5"] = "=totally_undefined_name*2"
    path = out("29_external_links_and_name_errors.xlsx")
    wb.save(path)
    add_external_link_stub(path)
    record(
        "29_external_links_and_name_errors.xlsx", "5-edge-case", "External workbook links and #NAME? errors",
        "A formula authored against another workbook ('[Budget-2023.xlsx]Summary'!$B$2), a "
        "stub xl/externalLinks/ part exercising the ExternalLink/Unsupported path, and a "
        "reference to an undefined name (#NAME?).",
        ["external workbook reference (stub)", "#NAME? via undefined name"],
    )


def build_30_formula_function_showcase():
    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README", 0)
    readme["A1"] = "Excel formula function showcase"
    readme["A1"].font = Font(bold=True, size=16)
    readme["A3"] = "One sheet per function category. Column A = function name, column B = a live"
    readme["A4"] = "example formula using that function, column C = whether it is inside"
    readme["A5"] = 'Marksheet\'s portable-a1@1 evaluation profile ("core") or outside it.'
    readme["A7"] = "Seed data (columns H:N, same on every category sheet):"
    readme["A8"] = "H=numbers, I=paired numbers, J=text, K=dates, L=booleans, M=lookup key, N=lookup value"
    r = 10
    total = 0
    unique = set()
    readme.cell(row=r, column=1, value="Category").font = Font(bold=True)
    readme.cell(row=r, column=2, value="Function count").font = Font(bold=True)
    for category, items in CATALOG.items():
        r += 1
        readme.cell(row=r, column=1, value=category.replace("_", " "))
        readme.cell(row=r, column=2, value=len(items))
        total += len(items)
        unique.update(name for name, _ in items)
    r += 2
    readme.cell(row=r, column=1, value="TOTAL rows").font = Font(bold=True)
    readme.cell(row=r, column=2, value=total).font = Font(bold=True)
    r += 1
    readme.cell(row=r, column=1, value="Unique function names").font = Font(bold=True)
    readme.cell(row=r, column=2, value=len(unique)).font = Font(bold=True)
    readme.column_dimensions["A"].width = 40

    sheet_titles = []
    for category, items in CATALOG.items():
        title = category.replace("_", " ")[:31]
        ws = wb.create_sheet(title)
        sheet_titles.append(title)
        write_seed_block(ws)
        ws.cell(row=1, column=1, value="Function").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Formula").font = Font(bold=True)
        ws.cell(row=1, column=3, value="Profile").font = Font(bold=True)
        for i, (name, formula) in enumerate(items, start=2):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=formula)
            ws.cell(
                row=i, column=3,
                value="portable-a1@1 core" if name in PORTABLE_A1_FUNCTIONS else "",
            )
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 18

    wb.save(out("30_formula_function_showcase.xlsx"))
    record(
        "30_formula_function_showcase.xlsx", "5-edge-case", "Every Excel function, one row each",
        f"{total} formula rows across {len(sheet_titles)} category sheets covering "
        f"{len(unique)} distinct Excel function names (Math & Trig, Statistical, Lookup & "
        "Reference, Text, Logical, Date & Time, Financial, Information, Engineering, "
        "Database, Web, Dynamic Array, Cube, and legacy Compatibility aliases) -- the direct "
        '"every Excel formula represented" file.',
        [f"{total} formulas", f"{len(unique)} unique functions", f"{len(sheet_titles)} category sheets"],
    )


def build_31_openpyxl_absolute_relationship_targets():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "This file is deliberately left as raw openpyxl output -- see README.md"
    ws["A2"] = "Its xl/_rels/workbook.xml.rels uses Target=\"/xl/worksheets/sheet1.xml\""
    ws["A3"] = "(a package-absolute OPC target), which this project's importer rejects."
    ws["B5"] = 42
    path = out("31_openpyxl_absolute_relationship_targets.xlsx")
    wb.save(path)
    # Deliberately NOT normalized -- see SKIP_NORMALIZE below.
    record(
        "31_openpyxl_absolute_relationship_targets.xlsx", "6-regression",
        "Regression: openpyxl's absolute relationship targets",
        "Left exactly as openpyxl wrote it (every other file in this corpus has this "
        "normalized away). openpyxl always emits package-absolute relationship targets "
        "(Target=\"/xl/worksheets/sheet1.xml\") for worksheets/tables/charts/comments. "
        "crates/marksheet-convert/src/xlsx/package.rs's resolve_target() rejects any "
        "target starting with \"/\" outright (MS4105, 'relationship target is absolute, "
        "external, or malformed'), which used to fail this file outright. Since "
        "openpyxl is one of the most common tools people use to generate .xlsx "
        "programmatically, this was a real compatibility gap rather than a contrived "
        "edge case. Now fixed; kept as the regression fixture, with the original "
        "targets preserved rather than normalized.",
        ["absolute OPC relationship targets", "MS4105", "fixed"],
    )


def build_32_sheet_scoped_defined_name():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "local target"
    ws["B1"] = 7
    ws.defined_names.add(DefinedName("local_note", attr_text="Sheet1!$B$1"))
    ws["A3"] = "Formula using the sheet-scoped name"
    ws["B3"] = "=local_note*2"
    path = out("32_sheet_scoped_defined_name.xlsx")
    wb.save(path)
    record(
        "32_sheet_scoped_defined_name.xlsx", "6-regression", "Sheet-scoped (local) defined name",
        "A single defined name scoped to one sheet rather than the workbook. "
        "crates/marksheet-convert used to reject the whole file for this "
        '("sheet-local defined names are outside the initial profile", MS4105). Now '
        "fixed; kept as the regression fixture.",
        ["sheet-scoped defined name", "fixed"],
    )


def build_33_null_error_token():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Intersection producing #NULL! in real Excel"
    ws["A2"] = "=SUM((A10:A11 C10:C11))"
    path = out("33_null_error_token.xlsx")
    wb.save(path)
    record(
        "33_null_error_token.xlsx", "5-edge-case", "#NULL! / space-intersection formula",
        "A range-intersection formula (space operator) that evaluates to #NULL! in real Excel. "
        "SPEC.md section 13's required runtime-failure table does not include #NULL! (or "
        "#SPILL!) -- verified behavior: the importer does NOT reject the file for this, it "
        'gracefully substitutes the formula ("unsupported Excel formula was replaced with '
        '#NAME?", a normal lossy/approximated outcome), unlike the true rejection cases in the '
        "6-regression tier below.",
        ["#NULL! / space-intersection formula", "graceful lossy substitution"],
    )


def build_34_external_hyperlink():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "External hyperlink"
    ws["A1"].hyperlink = "https://example.com/reports/q1"
    ws["A1"].font = Font(color="0563C1", underline="single")
    ws["A3"] = "Unrelated data that should be unaffected by the link above"
    ws["A4"] = 123
    path = out("34_external_hyperlink.xlsx")
    wb.save(path)
    record(
        "34_external_hyperlink.xlsx", "6-regression", "External hyperlink relationship",
        "One ordinary external hyperlink (TargetMode=\"External\"). "
        "crates/marksheet-convert/src/xlsx/package.rs rejects any package relationship with "
        'TargetMode="External" during its upfront package-wide relationship validation, before '
        'per-feature handling runs ("external OOXML relationships are rejected", MS4105) -- so '
        "an external hyperlink anywhere in the workbook used to take down the entire import, "
        "not just that cell's link. Now fixed; kept as the regression fixture. Internal "
        "same-sheet hyperlinks were always unaffected -- see "
        "26_merged_cells_comments_hyperlinks.xlsx.",
        ["external hyperlink", "fixed"],
    )


def build_35_custom_format_date_false_positive():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Ordinary red-negative custom format, misread as a date"
    c = ws.cell(row=2, column=1, value=-42.5)
    c.number_format = "0.00;[Red]-0.00"
    path = out("35_custom_format_date_false_positive.xlsx")
    wb.save(path)
    record(
        "35_custom_format_date_false_positive.xlsx", "6-regression",
        "Custom number format false-positives as a date",
        'A negative number under the extremely common "0.00;[Red]-0.00" custom format (red '
        "negatives, no color scale). crates/marksheet-convert/src/xlsx/import.rs's "
        "apply_number_format classified any custom format code containing the letter \"y\" or "
        "\"d\" as Date/DateTime -- and \"[Red]\" contains a \"d\", so this ordinary format was "
        "misread as a date and the negative value rejected outright "
        '("negative Excel date serial is outside the initial profile", MS4105). Now fixed; '
        "kept as the regression fixture.",
        ["custom number format date misclassification", "fixed"],
    )


BUILDERS = [
    build_01_empty_workbook,
    build_02_single_cell,
    build_03_scalar_values_no_formulas,
    build_04_personal_budget,
    build_05_grade_book,
    build_06_todo_checklist,
    build_07_invoice,
    build_08_unit_conversion,
    build_09_loan_payment_calculator,
    build_10_multi_sheet_sales_report,
    build_11_lookup_directory,
    build_12_named_ranges_tax_calc,
    build_13_excel_table_orders,
    build_14_conditional_formatting_dashboard,
    build_15_data_validation_form,
    build_16_frozen_panes_large_grid,
    build_17_financial_model_3_statement,
    build_18_large_dataset,
    build_19_array_and_dynamic_formulas,
    build_20_error_handling_showcase,
    build_21_circular_reference,
    build_22_charts,
    build_23_pivot_table_source,
    build_24_unicode_and_special_characters,
    build_25_many_sheets,
    build_26_merged_cells_comments_hyperlinks,
    build_27_macro_enabled,
    build_28_number_formats_and_styles,
    build_29_external_links_and_name_errors,
    build_30_formula_function_showcase,
    build_31_openpyxl_absolute_relationship_targets,
    build_32_sheet_scoped_defined_name,
    build_33_null_error_token,
    build_34_external_hyperlink,
    build_35_custom_format_date_false_positive,
]

SKIP_NORMALIZE = {"31_openpyxl_absolute_relationship_targets.xlsx"}


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for existing in OUT_DIR.glob("*"):
        existing.unlink()
    for builder in BUILDERS:
        builder()
    for f in OUT_DIR.glob("*"):
        if f.name not in SKIP_NORMALIZE:
            normalize_absolute_rels(f)
    manifest_path = HERE / "manifest.json"
    manifest_path.write_text(json.dumps(MANIFEST, indent=2) + "\n")
    print(f"Generated {len(MANIFEST)} files into {OUT_DIR}")
    print(f"Manifest written to {manifest_path}")


if __name__ == "__main__":
    main()
